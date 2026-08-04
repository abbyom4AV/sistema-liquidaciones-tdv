from __future__ import annotations

import os
import shutil
import tempfile
import unittest
from decimal import Decimal
from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook

from procesamientos.services.generacion_sifa import (
    reconstruir_resultado_para_escritura_sifa,
    serializar_lineas_preparadas_sifa,
)
from services.sifa.extractor import (
    COLUMNAS_GASTO,
    extraer_liquidacion_sifa,
)
from services.sifa.validator import (
    LineaPreparadaSifa,
    validar_liquidacion_sifa,
)
from services.sifa.matcher import ResultadoMatcherSifa, LineaDespachoSifa
from services.sifa.writer import (
    construir_valores_fila_sifa,
    escribir_archivo_sifa,
    _resolver_rutas_raw_data,
)


MUESTRA = Path(os.environ.get("TEMP", tempfile.gettempdir())) / (
    "sifa_liq_sample.xlsx"
)
FUENTE_MUESTRA = Path(
    r"c:\Users\aobando\Fruta Internacional\GONZÁLES, César (TDV) "
    r"- Liquidaciones TDV\Clientes Liquidaciones PDF\SIFA"
    r"\2026\semana 21\Copia de LICQUIDACION SI.FA. V.2621.xlsx"
)


def _asegurar_muestra() -> Path:
    if MUESTRA.is_file():
        return MUESTRA
    if FUENTE_MUESTRA.is_file():
        shutil.copy2(FUENTE_MUESTRA, MUESTRA)
        return MUESTRA
    raise unittest.SkipTest("No hay Excel de muestra SIFA.")


class ExtraccionSifaTests(unittest.TestCase):
    def test_extrae_lineas_gastos_y_vertical(self):
        ruta = _asegurar_muestra()
        liq = extraer_liquidacion_sifa(ruta)
        self.assertEqual(liq.factura_corta, "5513")
        self.assertEqual(len(liq.contenedores_header), 10)
        self.assertEqual(liq.total_cajas, 16320)
        self.assertEqual(liq.gastos["Deduction Eur"], Decimal("3152"))
        self.assertEqual(liq.gastos["Fresco Eur"], Decimal("52"))
        self.assertEqual(
            liq.gastos["Transportation Eur"],
            Decimal("4100.38"),
        )
        self.assertEqual(liq.gastos["Logistica Eur"], Decimal("2520"))
        self.assertEqual(liq.gastos["Handling Eur"], Decimal("915"))
        self.assertEqual(
            liq.gastos["Transportation Cus. Eur"],
            Decimal("11550"),
        )
        self.assertEqual(liq.total_costos_excel, Decimal("22289.38"))
        self.assertEqual(liq.rubros_no_mapeados, ())
        verticales = [
            ln for ln in liq.lineas if ln.carton.startswith("VERTICAL ")
        ]
        self.assertGreaterEqual(len(verticales), 1)
        self.assertTrue(
            all(ln.calibre == 6 for ln in verticales[:3])
        )
        verdes = [ln for ln in liq.lineas if ln.tipo_fruta == "VERDE"]
        self.assertEqual(len(verdes), 6)
        self.assertTrue(all(ln.sin_comision for ln in liq.lineas))
        self.assertEqual(len(liq.comisiones_contenedor), 10)


class ValidacionYEscrituraSifaTests(unittest.TestCase):
    def _linea(
        self,
        *,
        contenedor: str = "SEGU9725656",
        carton: str = "GOLDEN DIAMOND",
        calibre: int = 6,
        cajas: int = 160,
        tipo: str = "ESPECIAL",
        precio: str = "10",
        comision: str = "0",
    ) -> LineaPreparadaSifa:
        gastos = {col: Decimal("100") for col in COLUMNAS_GASTO}
        return LineaPreparadaSifa(
            contenedor=contenedor,
            nave="CALA PULA V.2621",
            destino="VADO LIGURE",
            tipo_fruta=tipo,
            carton=carton,
            calibre=calibre,
            total_cajas=cajas,
            semana=21,
            anio=2026,
            semana_texto="21-2026",
            cliente_raw="SIFA",
            precio_venta_eur=Decimal(precio),
            comision=Decimal(comision),
            sin_comision_linea=True,
            gastos=gastos,
            factura_corta="5513",
            fila_origen=22,
        )

    def test_serializa_y_reconstruye(self):
        lineas = (self._linea(), self._linea(calibre=7, cajas=80))
        serializadas = serializar_lineas_preparadas_sifa(lineas)
        resultado = reconstruir_resultado_para_escritura_sifa(
            anio=2026,
            semana=21,
            destino_ui="VADO LIGURE",
            lineas_preparadas=serializadas,
            total_cajas_liquidacion=240,
            total_cajas_despachos=240,
            comision_total=0,
            resumen_gastos=lineas[0].gastos,
        )
        self.assertTrue(resultado.puede_escribir)
        self.assertEqual(len(resultado.validacion.lineas_preparadas), 2)
        fila = construir_valores_fila_sifa(resultado, 0)
        self.assertEqual(fila["Cliente"], "SIFA")
        self.assertEqual(fila["# Calibre"], "6")
        self.assertEqual(fila["Año"], "2026")
        self.assertEqual(fila["Deduction Eur"], 100)
        self.assertEqual(fila["Comisión €"], 0)
        self.assertEqual(fila["Precio de Venta €"], 10.0)

    def test_writer_zip_digitados(self):
        lineas = (self._linea(),)
        serializadas = serializar_lineas_preparadas_sifa(lineas)
        resultado = reconstruir_resultado_para_escritura_sifa(
            anio=2026,
            semana=21,
            destino_ui="VADO LIGURE",
            lineas_preparadas=serializadas,
            total_cajas_liquidacion=160,
            total_cajas_despachos=160,
            resumen_gastos=lineas[0].gastos,
        )

        with tempfile.TemporaryDirectory() as tmp:
            origen = Path(tmp) / "origen.xlsx"
            salida = Path(tmp) / "salida.xlsx"
            wb = Workbook()
            # openpyxl crea sheet1; renombramos y añadimos hojas
            # para imitar orden SIFA (Raw Data = sheet3).
            ws1 = wb.active
            ws1.title = "Hoja1"
            wb.create_sheet("Precios")
            ws = wb.create_sheet("Raw Data")
            headers = [
                "Año",
                "Semana",
                "Cliente",
                "Nave",
                "Contenedor ",
                "Destino",
                "Tipo de fruta",
                "# Calibre",
                "Total Cajas",
                "Cartón",
                "Deduction Eur",
                "Fresco Eur",
                "Transportation Eur",
                "Logistica Eur",
                "Handling Eur",
                "Transportation Cus. Eur",
                "Comisión €",
                "Precio de Venta €",
            ]
            for col, nombre in enumerate(headers, start=1):
                ws.cell(1, col, nombre)
            ws.cell(2, 1, 2025)
            ws.cell(2, 2, "01-2025")
            ws.cell(2, 3, "SIFA")
            ws.cell(2, 5, "AAAA1111111")
            from openpyxl.worksheet.table import Table, TableStyleInfo

            tabla = Table(displayName="Tabla1", ref="A1:R2")
            tabla.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(tabla)
            wb.save(origen)

            # Confirmar resolución dinámica de Raw Data / Tabla1.
            with ZipFile(origen) as zipped:
                hoja, tabla = _resolver_rutas_raw_data(zipped)
                self.assertTrue(hoja.endswith(".xml"))
                self.assertIn("worksheets/", hoja)
                self.assertIn("tables/", tabla)

            escritura = escribir_archivo_sifa(
                procesamiento=resultado,
                ruta_archivo_cliente=origen,
                ruta_salida=salida,
            )
            self.assertEqual(escritura.filas_agregadas, 1)
            self.assertTrue(salida.is_file())


class ValidacionContenedoresSifaTests(unittest.TestCase):
    def test_detecta_desajuste_cajas(self):
        ruta = _asegurar_muestra()
        liq = extraer_liquidacion_sifa(ruta)
        # Despachos sintéticos con mismos contenedores pero cajas mal.
        lineas_desp = []
        for cont in liq.contenedores_header:
            lineas_desp.append(
                LineaDespachoSifa(
                    fila_excel=1,
                    semana=21,
                    anio=2026,
                    semana_texto="21-2026",
                    contenedor=cont,
                    cliente="SI.FA. SRL",
                    barco="CALA PULA V.2621",
                    puerto_destino="VADO LIGURE",
                    tipo_empaque="ESPECIAL",
                    carton="X",
                    calibre=6,
                    total_cajas=1,
                    factura="5513",
                    factura_corta="5513",
                )
            )
        desp = ResultadoMatcherSifa(
            archivo="",
            hoja="Base Datos",
            cliente_buscado="SI.FA. SRL",
            semana=21,
            anio=2026,
            destino_buscado="VADO LIGURE",
            semana_texto="21-2026",
            lineas=tuple(lineas_desp),
            total_cajas=len(lineas_desp),
            contenedores=tuple(liq.contenedores_header),
            destinos=("VADO LIGURE",),
            facturas_cortas=("5513",),
            naves=("CALA PULA V.2621",),
        )
        validacion = validar_liquidacion_sifa(
            liq,
            desp,
            "VADO LIGURE",
        )
        self.assertFalse(validacion.es_valido)
        codigos = {e.codigo for e in validacion.errores}
        self.assertIn("TOTAL_CAJAS_NO_COINCIDE", codigos)


class ComisionSifaTests(unittest.TestCase):
    def test_parsea_commission_porcentaje_y_asigna_total(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        # Header mínimo
        ws["C14"] = "VESSEL: TEST V.1"
        ws["H14"] = "VADO LIGURE,ITALY"
        ws["G16"] = "INV."
        ws["H16"] = "1234 DEL 01/01/2026"
        ws["C17"] = "CONTAINER:"
        ws["D17"] = "SEGU9502251/SEGU9502631"

        # Bloque 1
        ws["C21"] = "SEGU9502251"
        ws["D21"] = "MARCA"
        ws["E21"] = "COUNT"
        ws["F21"] = "CARTONS"
        ws["G21"] = "NOTES"
        ws["H21"] = "AMOUNT in EURO"
        ws["I21"] = "€/CRT/NET"
        ws["C22"] = "Pines"
        ws["D22"] = "GOLDEN DIAMOND"
        ws["E22"] = 6
        ws["F22"] = 100
        ws["G22"] = "CODIGO:1353038"
        ws["H22"] = 1000
        ws["I22"] = 10
        # CODIGO en NOTES ⇒ sin comisión (aunque no diga NO COMMISSION)
        ws["C23"] = "TOTAL"
        ws["F23"] = 100
        ws["H23"] = 1000

        # Bloque 2 — línea CON comisión (sin CODIGO ni NO COMMISSION)
        ws["C25"] = "SEGU9502631"
        ws["D25"] = "MARCA"
        ws["E25"] = "COUNT"
        ws["F25"] = "CARTONS"
        ws["G25"] = "NOTES"
        ws["H25"] = "AMOUNT in EURO"
        ws["I25"] = "€/CRT/NET"
        ws["C26"] = "Pines"
        ws["D26"] = "SWEET DIAMOND"
        ws["E26"] = 7
        ws["F26"] = 50
        ws["H26"] = 500
        ws["I26"] = 10
        ws["C27"] = "TOTAL"
        ws["F27"] = 50

        # Gastos
        ws["C30"] = "COSIARMA - THC/BL FEE/BOLLO at destination total"
        ws["H30"] = 100
        ws["C31"] = "FRESCO-DELIVERY"
        ws["H31"] = 50
        ws["C32"] = "FRESCO-TRANSPORT TO WEREHOUSE AND TERMINAL COSTS TOTAL"
        ws["H32"] = 200
        ws["C33"] = "2S LOGISTICA-WEREHOUSE DESCHARGED"
        ws["H33"] = 150
        ws["C34"] = "2S LOGISTICA-HANDLING AND PICKING"
        ws["H34"] = 75
        ws["C35"] = "CAA-TRANSPORT TO CUSTOMER"
        ws["H35"] = 425
        # GASTOS TOTAL (solo gastos, no TOTAL OF COSTS)
        ws["C36"] = "GASTOS TOTAL PUERTO-CLIENTES FINAL"
        ws["H36"] = "TOTAL:"
        ws["I36"] = 1000
        # Comisiones
        ws["C38"] = "commission 7% SEGU9502251"
        ws["H38"] = 1494.50
        ws["C39"] = "commission 7% SEGU9502631"
        ws["H39"] = 1417.50
        ws["C40"] = "TOTAL OF COSTS"
        ws["H40"] = 3912  # 1000 gastos + 2912 comisión

        with tempfile.TemporaryDirectory() as tmp:
            ruta = Path(tmp) / "liq.xlsx"
            wb.save(ruta)
            liq = extraer_liquidacion_sifa(ruta)

        self.assertEqual(len(liq.lineas), 2)
        self.assertTrue(liq.lineas[0].sin_comision)  # CODIGO
        self.assertFalse(liq.lineas[1].sin_comision)  # con comisión
        self.assertEqual(liq.lineas[0].tipo_fruta, "ESPECIAL")
        self.assertEqual(len(liq.comisiones_contenedor), 2)
        self.assertEqual(
            liq.comisiones_contenedor[0].monto_eur,
            Decimal("1494.5"),
        )
        self.assertEqual(
            liq.comisiones_contenedor[1].monto_eur,
            Decimal("1417.5"),
        )
        self.assertEqual(liq.total_costos_excel, Decimal("3912"))
        suma_gastos = sum(liq.gastos.values())
        self.assertEqual(suma_gastos, Decimal("1000"))

        # Validar asignación de comisión a líneas
        lineas_desp = []
        for cont, cajas in (("SEGU9502251", 100), ("SEGU9502631", 50)):
            lineas_desp.append(
                LineaDespachoSifa(
                    fila_excel=1,
                    semana=21,
                    anio=2026,
                    semana_texto="21-2026",
                    contenedor=cont,
                    cliente="SI.FA. SRL",
                    barco="TEST",
                    puerto_destino="VADO LIGURE",
                    tipo_empaque="ESPECIAL",
                    carton="X",
                    calibre=6,
                    total_cajas=cajas,
                    factura="1234",
                    factura_corta="1234",
                )
            )
        desp = ResultadoMatcherSifa(
            archivo="",
            hoja="Base Datos",
            cliente_buscado="SI.FA. SRL",
            semana=21,
            anio=2026,
            destino_buscado="VADO LIGURE",
            semana_texto="21-2026",
            lineas=tuple(lineas_desp),
            total_cajas=150,
            contenedores=("SEGU9502251", "SEGU9502631"),
            destinos=("VADO LIGURE",),
            facturas_cortas=("1234",),
            naves=("TEST",),
        )
        validacion = validar_liquidacion_sifa(
            liq,
            desp,
            "VADO LIGURE",
        )
        self.assertTrue(validacion.es_valido)
        self.assertEqual(
            validacion.comision_total,
            Decimal("2912.0"),
        )
        # Línea con CODIGO → 0; línea con comisión → total
        self.assertEqual(
            validacion.lineas_preparadas[0].comision,
            Decimal("0"),
        )
        self.assertEqual(
            validacion.lineas_preparadas[1].comision,
            Decimal("2912.0"),
        )
        codigos = {a.codigo for a in validacion.advertencias}
        self.assertNotIn("TOTAL_COSTOS_NO_CALZA", codigos)


if __name__ == "__main__":
    unittest.main()
