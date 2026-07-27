from pathlib import Path
from tempfile import TemporaryDirectory
import os
import unittest

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from services.dimanno.processor import preparar_procesamiento_dimanno
from services.dimanno.writer import (
    LiquidacionDuplicadaError,
    escribir_archivo_dimanno,
)


BASE_DIR = Path(__file__).resolve().parents[1]
CARPETA_PRUEBA = BASE_DIR / "PruebaSemana15 - Dimanno"

RUTA_LIQUIDACION = (
    CARPETA_PRUEBA
    / "Copia de Esquema gastos TROPICALES 2026 - Hasta semana 22.xlsx"
)

RUTA_DESPACHOS = (
    CARPETA_PRUEBA
    / "Despachos TDV 2025 V18.xlsx"
)

RUTA_CLIENTE_SIN_SEMANA = (
    CARPETA_PRUEBA
    / "DIMANNO Liquidaciones v2.1 Sin Semana 15.xlsx"
)

RUTA_CLIENTE_CON_SEMANA = (
    CARPETA_PRUEBA
    / "DIMANNO Liquidaciones v2.1 Con Semana 15.xlsx"
)

ARCHIVOS_DISPONIBLES = all(
    ruta.is_file()
    for ruta in (
        RUTA_LIQUIDACION,
        RUTA_DESPACHOS,
        RUTA_CLIENTE_SIN_SEMANA,
        RUTA_CLIENTE_CON_SEMANA,
    )
)

EJECUTAR_PRUEBAS_EXCEL = os.getenv("RUN_EXCEL_TESTS") == "1"


@unittest.skipUnless(
    ARCHIVOS_DISPONIBLES and EJECUTAR_PRUEBAS_EXCEL,
    (
        "Prueba lenta de integración con Excel. "
        "Ejecute con RUN_EXCEL_TESTS=1."
    ),
)
class PruebasEscritorDimanno(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.procesamiento = preparar_procesamiento_dimanno(
            ruta_liquidacion=RUTA_LIQUIDACION,
            nombre_hoja="FT 5292 W15",
            ruta_despachos=RUTA_DESPACHOS,
            anio=2026,
        )

    def test_escribe_semana_15(self) -> None:
        with TemporaryDirectory(
            ignore_cleanup_errors=True,
        ) as carpeta_temporal:
            ruta_salida = (
                Path(carpeta_temporal)
                / "resultado_semana15.xlsx"
            )

            resultado = escribir_archivo_dimanno(
                procesamiento=self.procesamiento,
                ruta_archivo_cliente=(
                    RUTA_CLIENTE_SIN_SEMANA
                ),
                ruta_salida=ruta_salida,
                recalcular_al_final=False,
            )

            self.assertTrue(ruta_salida.is_file())
            self.assertEqual(resultado.filas_agregadas, 5)
            self.assertEqual(resultado.fila_inicial, 309)
            self.assertEqual(resultado.fila_final, 313)
            self.assertEqual(
                resultado.destino_final,
                "VADO LIGURE",
            )

            libro = load_workbook(
                ruta_salida,
                read_only=False,
                data_only=False,
                keep_links=True,
            )

            try:
                hoja = libro["Raw Data"]
                tabla = hoja.tables["Tabla1"]

                self.assertEqual(
                    tabla.ref,
                    "A1:BX313",
                )

                # Solo columnas de Tabla1: fuera de la tabla
                # hay encabezados duplicados (p. ej. Total Cajas).
                columna_min, _, columna_max, _ = (
                    range_boundaries(tabla.ref)
                )

                encabezados = {
                    hoja.cell(row=1, column=columna).value:
                    columna
                    for columna in range(
                        columna_min,
                        columna_max + 1,
                    )
                }

                filas_esperadas = [
                    (
                        "CXRU1615350",
                        "Especial",
                        6,
                        1050,
                        18.82,
                    ),
                    (
                        "CXRU1615350",
                        "Especial",
                        7,
                        525,
                        18.89,
                    ),
                    (
                        "SEGU9548592",
                        "Intermedio",
                        6,
                        480,
                        14.15,
                    ),
                    (
                        "SEGU9548592",
                        "Intermedio",
                        7,
                        1040,
                        14.096,
                    ),
                    (
                        "SEGU9548592",
                        "Intermedio",
                        8,
                        160,
                        13.17,
                    ),
                ]

                for fila_excel, esperado in zip(
                    range(309, 314),
                    filas_esperadas,
                ):
                    (
                        contenedor,
                        tipo_fruta,
                        calibre,
                        cajas,
                        precio,
                    ) = esperado

                    self.assertEqual(
                        hoja.cell(
                            fila_excel,
                            encabezados["Contenedor "],
                        ).value,
                        contenedor,
                    )

                    self.assertEqual(
                        hoja.cell(
                            fila_excel,
                            encabezados["Tipo de fruta"],
                        ).value,
                        tipo_fruta,
                    )

                    self.assertEqual(
                        hoja.cell(
                            fila_excel,
                            encabezados["# Calibre"],
                        ).value,
                        calibre,
                    )

                    self.assertEqual(
                        hoja.cell(
                            fila_excel,
                            encabezados["Total Cajas"],
                        ).value,
                        cajas,
                    )

                    self.assertAlmostEqual(
                        hoja.cell(
                            fila_excel,
                            encabezados["Precio de Venta €"],
                        ).value,
                        precio,
                        places=6,
                    )

                    formula = hoja.cell(
                        fila_excel,
                        encabezados["Contar Fcls"],
                    ).value

                    self.assertIsInstance(formula, str)
                    self.assertTrue(formula.startswith("="))

            finally:
                libro.close()

    def test_impide_procesar_la_misma_liquidacion(
        self,
    ) -> None:
        with TemporaryDirectory(
            ignore_cleanup_errors=True,
        ) as carpeta_temporal:
            ruta_salida = (
                Path(carpeta_temporal)
                / "resultado_duplicado.xlsx"
            )

            with self.assertRaises(
                LiquidacionDuplicadaError
            ):
                escribir_archivo_dimanno(
                    procesamiento=self.procesamiento,
                    ruta_archivo_cliente=(
                        RUTA_CLIENTE_CON_SEMANA
                    ),
                    ruta_salida=ruta_salida,
                    recalcular_al_final=False,
                )


if __name__ == "__main__":
    unittest.main()
