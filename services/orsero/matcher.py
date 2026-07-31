from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass, is_dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from services.dimanno.matcher import (
    FormatoDespachosError,
    convertir_entero,
    detectar_encabezados,
    interpretar_semana,
    normalizar_texto,
    texto_limpio,
)


CLIENTE_ORSERO = "ORSERO"


class ErrorMatcherOrsero(Exception):
    """Error general al cruzar Despachos para Orsero."""


class FormatoDespachosOrseroError(ErrorMatcherOrsero):
    """El archivo de Despachos no tiene el formato esperado."""


class SinCoincidenciasOrseroError(ErrorMatcherOrsero):
    """No hay líneas Orsero Especial para la semana/año."""


@dataclass(frozen=True)
class LineaDespachoOrsero:
    fila_excel: int
    semana: int
    anio: int
    semana_texto: str
    contenedor: str
    cliente: str
    barco: str
    puerto_destino: str
    tipo_empaque: str
    carton: str
    calibre: int
    total_cajas: int

    @property
    def tipo_fruta(self) -> str:
        return "ESPECIAL"


@dataclass(frozen=True)
class ResultadoMatcherOrsero:
    archivo: str
    hoja: str
    cliente_buscado: str
    semana: int
    anio: int
    semana_texto: str
    lineas: tuple[LineaDespachoOrsero, ...]
    total_cajas: int
    contenedores: tuple[str, ...]
    destinos: tuple[str, ...]


def _cliente_coincide(cliente_fila: str, buscado: str) -> bool:
    return normalizar_texto(cliente_fila).replace(
        " ",
        "",
    ) == normalizar_texto(buscado).replace(" ", "")


def _es_especial(tipo_empaque: str) -> bool:
    return normalizar_texto(tipo_empaque) == "ESPECIAL"


def _valor(
    fila: tuple[Any, ...],
    posiciones: dict[str, int],
    campo: str,
) -> Any:
    indice = posiciones[campo] - 1
    if indice >= len(fila):
        return None
    return fila[indice]


def buscar_lineas_despachos_orsero(
    ruta_archivo: str | Path,
    semana: int,
    anio: int,
    cliente: str = CLIENTE_ORSERO,
) -> ResultadoMatcherOrsero:
    ruta = Path(ruta_archivo)
    if not ruta.is_file():
        raise ErrorMatcherOrsero(
            f"No existe el archivo de Despachos: {ruta}"
        )

    if not (1 <= int(semana) <= 53):
        raise FormatoDespachosOrseroError(
            "La semana debe estar entre 1 y 53."
        )
    if int(anio) < 2000:
        raise FormatoDespachosOrseroError(
            "El año de Despachos parece inválido."
        )

    libro = load_workbook(
        ruta,
        read_only=True,
        data_only=True,
    )

    try:
        if "Base Datos" not in libro.sheetnames:
            raise FormatoDespachosOrseroError(
                "No existe la hoja 'Base Datos' en Despachos."
            )

        hoja = libro["Base Datos"]

        try:
            fila_encabezado, posiciones = detectar_encabezados(
                hoja
            )
        except FormatoDespachosError as error:
            raise FormatoDespachosOrseroError(str(error)) from error

        lineas: list[LineaDespachoOrsero] = []

        for numero_fila, fila in enumerate(
            hoja.iter_rows(
                min_row=fila_encabezado + 1,
                values_only=True,
            ),
            start=fila_encabezado + 1,
        ):
            if fila is None or all(
                valor is None or str(valor).strip() == ""
                for valor in fila
            ):
                continue

            try:
                cliente_fila = texto_limpio(
                    _valor(fila, posiciones, "cliente")
                )
                if not _cliente_coincide(cliente_fila, cliente):
                    continue

                tipo_empaque = texto_limpio(
                    _valor(fila, posiciones, "tipo_empaque")
                )
                if not _es_especial(tipo_empaque):
                    continue

                semana_valor = _valor(
                    fila,
                    posiciones,
                    "semana",
                )
                semana_fila, anio_semana = interpretar_semana(
                    semana_valor
                )
                anio_fila = convertir_entero(
                    _valor(fila, posiciones, "anio"),
                    "año",
                )
                if semana_fila != int(semana):
                    continue
                if anio_fila != int(anio):
                    continue
                if (
                    anio_semana is not None
                    and anio_semana != anio_fila
                ):
                    # Preferir columna Año; ya filtramos por ella.
                    pass

                linea = LineaDespachoOrsero(
                    fila_excel=numero_fila,
                    semana=semana_fila,
                    anio=anio_fila,
                    semana_texto=texto_limpio(semana_valor),
                    contenedor=texto_limpio(
                        _valor(fila, posiciones, "contenedor")
                    ),
                    cliente=cliente_fila,
                    barco=texto_limpio(
                        _valor(fila, posiciones, "barco")
                    ),
                    puerto_destino=texto_limpio(
                        _valor(
                            fila,
                            posiciones,
                            "puerto_destino",
                        )
                    ),
                    tipo_empaque=tipo_empaque,
                    carton=texto_limpio(
                        _valor(fila, posiciones, "carton")
                    ),
                    calibre=convertir_entero(
                        _valor(fila, posiciones, "calibre"),
                        "calibre",
                    ),
                    total_cajas=convertir_entero(
                        _valor(
                            fila,
                            posiciones,
                            "total_cajas",
                        ),
                        "total_cajas",
                    ),
                )
            except (
                FormatoDespachosError,
                ValueError,
                TypeError,
            ) as error:
                raise FormatoDespachosOrseroError(
                    f"Error en fila {numero_fila} de "
                    f"Despachos: {error}"
                ) from error

            lineas.append(linea)

        if not lineas:
            raise SinCoincidenciasOrseroError(
                "No se encontraron líneas Especial en Despachos "
                f"para cliente '{cliente}', semana {semana} "
                f"y año {anio}."
            )

        total_cajas = sum(linea.total_cajas for linea in lineas)
        contenedores = tuple(
            dict.fromkeys(
                linea.contenedor
                for linea in lineas
                if linea.contenedor
            )
        )
        destinos = tuple(
            dict.fromkeys(
                linea.puerto_destino.strip().upper()
                for linea in lineas
                if linea.puerto_destino.strip()
            )
        )
        semana_texto = lineas[0].semana_texto or (
            f"{semana:02d}-{anio}"
        )

        return ResultadoMatcherOrsero(
            archivo=ruta.name,
            hoja="Base Datos",
            cliente_buscado=cliente,
            semana=int(semana),
            anio=int(anio),
            semana_texto=semana_texto,
            lineas=tuple(lineas),
            total_cajas=total_cajas,
            contenedores=contenedores,
            destinos=destinos,
        )
    finally:
        libro.close()


def convertir_a_json(valor: Any) -> Any:
    if isinstance(valor, Decimal):
        return format(valor, "f")
    if is_dataclass(valor) and not isinstance(valor, type):
        return convertir_a_json(asdict(valor))
    if isinstance(valor, dict):
        return {
            k: convertir_a_json(v) for k, v in valor.items()
        }
    if isinstance(valor, (list, tuple)):
        return [convertir_a_json(v) for v in valor]
    return valor


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Busca líneas Especial de Despachos para Orsero."
        ),
    )
    parser.add_argument("--despachos", required=True)
    parser.add_argument("--semana", type=int, required=True)
    parser.add_argument("--anio", type=int, required=True)
    parser.add_argument("--cliente", default=CLIENTE_ORSERO)
    args = parser.parse_args()
    print(
        json.dumps(
            convertir_a_json(
                buscar_lineas_despachos_orsero(
                    ruta_archivo=args.despachos,
                    semana=args.semana,
                    anio=args.anio,
                    cliente=args.cliente,
                )
            ),
            indent=2,
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
