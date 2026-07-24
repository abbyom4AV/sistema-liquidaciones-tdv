from __future__ import annotations

import argparse
import json
import re
import unicodedata
from dataclasses import asdict, dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


NOMBRE_HOJA_DESPACHOS = "Base Datos"

COLUMNAS_REQUERIDAS = {
    "semana",
    "anio",
    "contenedor",
    "cliente",
    "barco",
    "puerto_destino",
    "tipo_empaque",
    "carton",
    "calibre",
    "total_cajas",
    "factura",
}

ALIASES_COLUMNAS = {
    "SEMANA": "semana",
    "ANO": "anio",
    "CONTENEDOR": "contenedor",
    "CLIENTE": "cliente",
    "BARCO": "barco",
    "PUERTO DESTINO": "puerto_destino",
    "TIPO EMPAQUE": "tipo_empaque",
    "CARTON": "carton",
    "CALIBRE": "calibre",
    "TOTAL CAJAS": "total_cajas",
    "FACTURA": "factura",
}


class ErrorMatcherDimanno(Exception):
    """Error general durante el cruce de Despachos."""


class FormatoDespachosError(ErrorMatcherDimanno):
    """El archivo de Despachos no tiene el formato esperado."""


class SinCoincidenciasError(ErrorMatcherDimanno):
    """No se encontraron líneas para los criterios solicitados."""


@dataclass(frozen=True)
class LineaDespacho:
    fila_excel: int
    semana: int
    anio: int
    contenedor: str
    cliente: str
    barco: str
    puerto_destino: str
    tipo_empaque: str
    carton: str
    calibre: int
    total_cajas: int
    factura: str
    factura_corta: str


@dataclass(frozen=True)
class ResultadoMatcher:
    archivo: str
    hoja: str
    cliente_buscado: str
    anio_buscado: int
    semana_buscada: int
    factura_corta_buscada: str
    lineas: tuple[LineaDespacho, ...]
    total_cajas: int
    contenedores: tuple[str, ...]


def normalizar_texto(valor: Any) -> str:
    if valor is None:
        return ""

    texto = str(valor).strip().upper()
    texto = unicodedata.normalize("NFD", texto)

    texto = "".join(
        caracter
        for caracter in texto
        if unicodedata.category(caracter) != "Mn"
    )

    return re.sub(r"\s+", " ", texto)


def texto_limpio(valor: Any) -> str:
    if valor is None:
        return ""

    return str(valor).strip()


def convertir_entero(valor: Any, nombre_campo: str) -> int:
    if valor is None:
        raise FormatoDespachosError(
            f"El campo '{nombre_campo}' está vacío."
        )

    if isinstance(valor, bool):
        raise FormatoDespachosError(
            f"El campo '{nombre_campo}' contiene un valor booleano."
        )

    if isinstance(valor, int):
        return valor

    if isinstance(valor, float):
        if not valor.is_integer():
            raise FormatoDespachosError(
                f"El campo '{nombre_campo}' debe ser entero: {valor}."
            )

        return int(valor)

    texto = (
        str(valor)
        .strip()
        .replace("\xa0", "")
        .replace(" ", "")
    )

    if not texto:
        raise FormatoDespachosError(
            f"El campo '{nombre_campo}' está vacío."
        )

    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(",", "")

    try:
        numero = Decimal(texto)
    except InvalidOperation as error:
        raise FormatoDespachosError(
            f"No se pudo convertir '{valor}' a entero "
            f"en el campo '{nombre_campo}'."
        ) from error

    if numero != numero.to_integral_value():
        raise FormatoDespachosError(
            f"El campo '{nombre_campo}' debe ser entero: {numero}."
        )

    return int(numero)


def interpretar_semana(valor: Any) -> tuple[int, int | None]:
    """
    Acepta formatos como:

    15
    W15
    15-2026
    W 15
    """
    if valor is None:
        raise FormatoDespachosError(
            "La semana está vacía."
        )

    if isinstance(valor, int):
        return valor, None

    if isinstance(valor, float):
        if not valor.is_integer():
            raise FormatoDespachosError(
                f"La semana no es válida: {valor}."
            )

        return int(valor), None

    texto = normalizar_texto(valor)

    coincidencia = re.fullmatch(
        r"W?\s*(\d{1,2})(?:\s*[-/]\s*(\d{4}))?",
        texto,
    )

    if not coincidencia:
        raise FormatoDespachosError(
            f"No se pudo interpretar la semana '{valor}'."
        )

    semana = int(coincidencia.group(1))
    anio = (
        int(coincidencia.group(2))
        if coincidencia.group(2)
        else None
    )

    if semana < 1 or semana > 53:
        raise FormatoDespachosError(
            f"La semana está fuera de rango: {semana}."
        )

    return semana, anio


def normalizar_factura(valor: Any) -> str:
    """
    Convierte la factura a una cadena de dígitos.

    Las facturas reales deben mantenerse como texto en Excel para
    conservar correctamente todos sus dígitos.
    """
    if valor is None:
        raise FormatoDespachosError(
            "La factura está vacía."
        )

    if isinstance(valor, bool):
        raise FormatoDespachosError(
            "La factura contiene un valor booleano."
        )

    if isinstance(valor, int):
        texto = str(valor)

    elif isinstance(valor, float):
        if not valor.is_integer():
            raise FormatoDespachosError(
                f"La factura numérica no es entera: {valor}."
            )

        if abs(valor) >= 10**15:
            raise FormatoDespachosError(
                "La factura fue almacenada como número de más de "
                "15 dígitos y Excel pudo perder precisión. "
                "Debe almacenarse como texto."
            )

        texto = str(int(valor))

    else:
        texto = str(valor).strip()

    digitos = re.sub(r"\D", "", texto)

    if len(digitos) < 4:
        raise FormatoDespachosError(
            f"La factura '{valor}' no contiene al menos cuatro dígitos."
        )

    return digitos


def obtener_factura_corta(valor: Any) -> str:
    factura = normalizar_factura(valor)
    return factura[-4:]


def detectar_encabezados(
    hoja: Any,
    maximo_filas: int = 10,
) -> tuple[int, dict[str, int]]:
    """
    Busca la fila de encabezados y devuelve:

    - Número de fila.
    - Posición de cada columna requerida.
    """
    for numero_fila in range(
        1,
        min(hoja.max_row, maximo_filas) + 1,
    ):
        posiciones: dict[str, int] = {}

        for numero_columna, celda in enumerate(
            hoja[numero_fila],
            start=1,
        ):
            encabezado = normalizar_texto(celda.value)
            nombre_interno = ALIASES_COLUMNAS.get(encabezado)

            if not nombre_interno:
                continue

            if nombre_interno in posiciones:
                raise FormatoDespachosError(
                    f"La columna '{encabezado}' aparece repetida."
                )

            posiciones[nombre_interno] = numero_columna

        if COLUMNAS_REQUERIDAS.issubset(posiciones):
            return numero_fila, posiciones

    faltantes = ", ".join(sorted(COLUMNAS_REQUERIDAS))

    raise FormatoDespachosError(
        "No se encontró una fila con todas las columnas requeridas. "
        f"Columnas esperadas: {faltantes}."
    )


def obtener_valor(
    valores_fila: tuple[Any, ...],
    posiciones: dict[str, int],
    nombre_columna: str,
) -> Any:
    posicion_excel = posiciones[nombre_columna]
    indice_python = posicion_excel - 1

    if indice_python >= len(valores_fila):
        return None

    return valores_fila[indice_python]


def buscar_lineas_despachos(
    ruta_archivo: str | Path,
    cliente: str,
    anio: int,
    semana: int,
    factura_corta: str,
) -> ResultadoMatcher:
    ruta = Path(ruta_archivo)

    if not ruta.is_file():
        raise FileNotFoundError(
            f"No existe el archivo de Despachos: {ruta}"
        )

    factura_buscada = re.sub(
        r"\D",
        "",
        str(factura_corta),
    ).zfill(4)[-4:]

    cliente_buscado = normalizar_texto(cliente)

    libro = load_workbook(
        filename=ruta,
        read_only=True,
        data_only=True,
        keep_links=True,
    )

    try:
        if NOMBRE_HOJA_DESPACHOS not in libro.sheetnames:
            raise FormatoDespachosError(
                f"No existe la hoja '{NOMBRE_HOJA_DESPACHOS}'."
            )

        hoja = libro[NOMBRE_HOJA_DESPACHOS]

        fila_encabezados, posiciones = detectar_encabezados(
            hoja
        )

        lineas: list[LineaDespacho] = []

        for numero_fila, valores in enumerate(
            hoja.iter_rows(
                min_row=fila_encabezados + 1,
                values_only=True,
            ),
            start=fila_encabezados + 1,
        ):
            valor_cliente = obtener_valor(
                valores,
                posiciones,
                "cliente",
            )

            if normalizar_texto(valor_cliente) != cliente_buscado:
                continue

            try:
                valor_anio = convertir_entero(
                    obtener_valor(
                        valores,
                        posiciones,
                        "anio",
                    ),
                    f"Año, fila {numero_fila}",
                )

                if valor_anio != anio:
                    continue

                valor_semana, anio_en_semana = interpretar_semana(
                    obtener_valor(
                        valores,
                        posiciones,
                        "semana",
                    )
                )

                if valor_semana != semana:
                    continue

                if (
                    anio_en_semana is not None
                    and anio_en_semana != anio
                ):
                    continue

                valor_factura = obtener_valor(
                    valores,
                    posiciones,
                    "factura",
                )

                valor_factura_corta = obtener_factura_corta(
                    valor_factura
                )

                if valor_factura_corta != factura_buscada:
                    continue

                linea = LineaDespacho(
                    fila_excel=numero_fila,
                    semana=valor_semana,
                    anio=valor_anio,
                    contenedor=texto_limpio(
                        obtener_valor(
                            valores,
                            posiciones,
                            "contenedor",
                        )
                    ).upper(),
                    cliente=texto_limpio(valor_cliente).upper(),
                    barco=texto_limpio(
                        obtener_valor(
                            valores,
                            posiciones,
                            "barco",
                        )
                    ),
                    puerto_destino=texto_limpio(
                        obtener_valor(
                            valores,
                            posiciones,
                            "puerto_destino",
                        )
                    ).upper(),
                    tipo_empaque=texto_limpio(
                        obtener_valor(
                            valores,
                            posiciones,
                            "tipo_empaque",
                        )
                    ).title(),
                    carton=texto_limpio(
                        obtener_valor(
                            valores,
                            posiciones,
                            "carton",
                        )
                    ),
                    calibre=convertir_entero(
                        obtener_valor(
                            valores,
                            posiciones,
                            "calibre",
                        ),
                        f"Calibre, fila {numero_fila}",
                    ),
                    total_cajas=convertir_entero(
                        obtener_valor(
                            valores,
                            posiciones,
                            "total_cajas",
                        ),
                        f"Total Cajas, fila {numero_fila}",
                    ),
                    factura=normalizar_factura(valor_factura),
                    factura_corta=valor_factura_corta,
                )

                lineas.append(linea)

            except FormatoDespachosError as error:
                raise FormatoDespachosError(
                    f"Error en la fila {numero_fila}: {error}"
                ) from error

        if not lineas:
            raise SinCoincidenciasError(
                "No se encontraron líneas en Despachos para: "
                f"cliente={cliente}, año={anio}, semana={semana}, "
                f"factura={factura_buscada}."
            )

        total_cajas = sum(
            linea.total_cajas
            for linea in lineas
        )

        contenedores = tuple(
            dict.fromkeys(
                linea.contenedor
                for linea in lineas
            )
        )

        return ResultadoMatcher(
            archivo=ruta.name,
            hoja=NOMBRE_HOJA_DESPACHOS,
            cliente_buscado=cliente,
            anio_buscado=anio,
            semana_buscada=semana,
            factura_corta_buscada=factura_buscada,
            lineas=tuple(lineas),
            total_cajas=total_cajas,
            contenedores=contenedores,
        )

    finally:
        libro.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Busca líneas de una liquidación Di Manno "
            "en el archivo de Despachos."
        )
    )

    parser.add_argument(
        "--archivo",
        required=True,
        help="Ruta del archivo de Despachos.",
    )

    parser.add_argument(
        "--cliente",
        default="DI MANNO",
        help="Nombre del cliente en Despachos.",
    )

    parser.add_argument(
        "--anio",
        type=int,
        required=True,
        help="Año de la liquidación.",
    )

    parser.add_argument(
        "--semana",
        type=int,
        required=True,
        help="Semana de la liquidación.",
    )

    parser.add_argument(
        "--factura",
        required=True,
        help="Últimos cuatro dígitos de la factura.",
    )

    argumentos = parser.parse_args()

    resultado = buscar_lineas_despachos(
        ruta_archivo=argumentos.archivo,
        cliente=argumentos.cliente,
        anio=argumentos.anio,
        semana=argumentos.semana,
        factura_corta=argumentos.factura,
    )

    print(
        json.dumps(
            asdict(resultado),
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
    