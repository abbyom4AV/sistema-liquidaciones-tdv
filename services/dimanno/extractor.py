from __future__ import annotations

import argparse
import json
import re
import unicodedata
from dataclasses import asdict, dataclass, is_dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PATRON_HOJA = re.compile(
    r"^FT\s+(?P<factura>\d{4})\s+W(?P<semana>\d{1,2})$",
    re.IGNORECASE,
)

PATRONES_PRODUCTOS = (
    (
        re.compile(r"^(?:MERCE|MERE) CORONA\s+(\d+)$"),
        "Intermedio",
    ),
    (
        re.compile(r"^EXTRA CON CINTILLO\s+(\d+)$"),
        "Especial",
    ),
)

MAPEO_GASTOS = {
    "COMMISSIONI": "Comisión",
    "NAVIERA": "Flete Eu",
    "CONTROL DE CALIDAD": "Control calidad Eu",
    "THC": "THC",
    "TRASPORTO": "Transporte",
    "DOGANA": "Aduanas",
}

MARCADORES_CERO = {
    "",
    "-",
    "–",
    "—",
}


class ErrorExtraccionDimanno(Exception):
    """Error general durante la lectura de una liquidación Di Manno."""


class FormatoLiquidacionError(ErrorExtraccionDimanno):
    """El archivo no cumple el formato esperado para Di Manno."""


@dataclass(frozen=True)
class LineaVenta:
    descripcion_original: str
    tipo_fruta: str | None
    calibre: int | None
    cajas: int
    precio_eur: Decimal


@dataclass(frozen=True)
class LiquidacionDimanno:
    archivo: str
    hoja: str
    factura_corta: str
    semana: int
    contenedores: tuple[str, ...]
    naviera: str
    destino: str
    total_cajas: int
    total_venta_eur: Decimal
    productos: tuple[LineaVenta, ...]
    gastos: dict[str, Decimal]
    rubros_no_mapeados: tuple[str, ...]


def normalizar_texto(valor: Any) -> str:
    """
    Normaliza texto para realizar comparaciones sin depender de
    mayúsculas, acentos o espacios adicionales.
    """
    if valor is None:
        return ""

    texto = str(valor).strip().upper()
    texto = unicodedata.normalize("NFD", texto)

    texto = "".join(
        caracter
        for caracter in texto
        if unicodedata.category(caracter) != "Mn"
    )

    return re.sub(r"\s+", " ", texto)


def convertir_decimal(valor: Any, nombre_campo: str) -> Decimal:
    """
    Convierte un valor de Excel a Decimal.

    Las celdas vacías y los distintos tipos de guion se consideran cero.
    """
    if valor is None:
        return Decimal("0")

    if isinstance(valor, Decimal):
        return valor

    if isinstance(valor, bool):
        raise FormatoLiquidacionError(
            f"El campo '{nombre_campo}' contiene un valor booleano."
        )

    if isinstance(valor, int):
        return Decimal(valor)

    if isinstance(valor, float):
        # El formato significativo elimina residuos binarios como
        # 4258.983200000001 sin redondear valores válidos como 14.096.
        return Decimal(format(valor, ".15g"))

    texto_original = str(valor).strip()

    if texto_original in MARCADORES_CERO:
        return Decimal("0")

    texto = (
        texto_original
        .replace("€", "")
        .replace("$", "")
        .replace("\xa0", "")
        .replace(" ", "")
    )

    # Formato europeo: 1.234,56
    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")

    elif "," in texto:
        texto = texto.replace(",", ".")

    try:
        return Decimal(texto)
    except InvalidOperation as error:
        raise FormatoLiquidacionError(
            f"No se pudo convertir '{valor}' a número "
            f"en el campo '{nombre_campo}'."
        ) from error


def convertir_entero(valor: Any, nombre_campo: str) -> int:
    numero = convertir_decimal(valor, nombre_campo)

    if numero != numero.to_integral_value():
        raise FormatoLiquidacionError(
            f"El campo '{nombre_campo}' debe ser entero, "
            f"pero contiene {numero}."
        )

    return int(numero)


def buscar_fila(
    hoja: Any,
    etiqueta: str,
    fila_inicial: int = 1,
    fila_final: int | None = None,
) -> int:
    """
    Busca una etiqueta exacta en la columna A y devuelve su fila.
    """
    objetivo = normalizar_texto(etiqueta)
    limite = fila_final or hoja.max_row

    for numero_fila in range(fila_inicial, limite + 1):
        contenido = normalizar_texto(
            hoja.cell(row=numero_fila, column=1).value
        )

        if contenido == objetivo:
            return numero_fila

    raise FormatoLiquidacionError(
        f"No se encontró la sección '{etiqueta}'."
    )


def interpretar_producto(
    descripcion: str,
) -> tuple[str | None, int | None]:
    texto = normalizar_texto(descripcion)

    for patron, tipo_fruta in PATRONES_PRODUCTOS:
        coincidencia = patron.match(texto)

        if coincidencia:
            calibre = int(coincidencia.group(1))
            return tipo_fruta, calibre

    return None, None


def separar_contenedores(valor: Any) -> tuple[str, ...]:
    texto = str(valor or "").strip()

    if not texto:
        raise FormatoLiquidacionError(
            "La liquidación no contiene contenedores."
        )

    contenedores = [
        contenedor.strip().upper()
        for contenedor in re.split(r"[/,;\n]+", texto)
        if contenedor.strip()
    ]

    if not contenedores:
        raise FormatoLiquidacionError(
            "No fue posible interpretar los contenedores."
        )

    # Conserva el orden y elimina duplicados.
    return tuple(dict.fromkeys(contenedores))


def extraer_productos(
    hoja: Any,
) -> tuple[tuple[LineaVenta, ...], int, Decimal]:
    fila_seccion = buscar_fila(hoja, "CONTO VENDITA")

    fila_encabezado = buscar_fila(
        hoja,
        "Descrizione del prodotto",
        fila_inicial=fila_seccion + 1,
    )

    fila_total = buscar_fila(
        hoja,
        "Totale vendita",
        fila_inicial=fila_encabezado + 1,
    )

    productos: list[LineaVenta] = []

    for numero_fila in range(fila_encabezado + 1, fila_total):
        descripcion = str(
            hoja.cell(row=numero_fila, column=1).value or ""
        ).strip()

        cajas = convertir_entero(
            hoja.cell(row=numero_fila, column=3).value,
            f"cajas de '{descripcion}'",
        )

        if cajas <= 0:
            continue

        precio = convertir_decimal(
            hoja.cell(row=numero_fila, column=6).value,
            f"precio de '{descripcion}'",
        )

        tipo_fruta, calibre = interpretar_producto(descripcion)

        productos.append(
            LineaVenta(
                descripcion_original=descripcion,
                tipo_fruta=tipo_fruta,
                calibre=calibre,
                cajas=cajas,
                precio_eur=precio,
            )
        )

    total_cajas = convertir_entero(
        hoja.cell(row=fila_total, column=3).value,
        "total de cajas",
    )

    total_venta = convertir_decimal(
        hoja.cell(row=fila_total, column=7).value,
        "total de venta",
    )

    cajas_calculadas = sum(producto.cajas for producto in productos)

    if cajas_calculadas != total_cajas:
        raise FormatoLiquidacionError(
            "El total de cajas de la liquidación no coincide con "
            "la suma de las líneas de producto. "
            f"Total informado: {total_cajas}. "
            f"Total calculado: {cajas_calculadas}."
        )

    return tuple(productos), total_cajas, total_venta


def extraer_gastos(
    hoja: Any,
) -> tuple[dict[str, Decimal], tuple[str, ...]]:
    fila_origen = buscar_fila(hoja, "Costi A Origine")
    total_origen = buscar_fila(
        hoja,
        "Totale costi origine",
        fila_inicial=fila_origen + 1,
    )

    fila_destino = buscar_fila(hoja, "Costi A Destino")
    total_destino = buscar_fila(
        hoja,
        "Totale costi a destino",
        fila_inicial=fila_destino + 1,
    )

    gastos = {
        nombre_destino: Decimal("0")
        for nombre_destino in MAPEO_GASTOS.values()
    }

    rubros_no_mapeados: list[str] = []
    rubros_encontrados: set[str] = set()

    rangos = (
        range(fila_origen + 2, total_origen),
        range(fila_destino + 2, total_destino),
    )

    for rango in rangos:
        for numero_fila in rango:
            descripcion_original = str(
                hoja.cell(row=numero_fila, column=1).value or ""
            ).strip()

            descripcion = normalizar_texto(descripcion_original)

            if not descripcion:
                continue

            if descripcion in MAPEO_GASTOS:
                if descripcion in rubros_encontrados:
                    raise FormatoLiquidacionError(
                        f"El rubro '{descripcion_original}' "
                        "aparece más de una vez."
                    )

                columna_destino = MAPEO_GASTOS[descripcion]

                gastos[columna_destino] = convertir_decimal(
                    hoja.cell(row=numero_fila, column=7).value,
                    descripcion_original,
                )

                rubros_encontrados.add(descripcion)
            else:
                rubros_no_mapeados.append(descripcion_original)

    faltantes = set(MAPEO_GASTOS) - rubros_encontrados

    if faltantes:
        raise FormatoLiquidacionError(
            "No se encontraron los siguientes rubros esperados: "
            + ", ".join(sorted(faltantes))
        )

    return gastos, tuple(rubros_no_mapeados)


def extraer_liquidacion(
    ruta_archivo: str | Path,
    nombre_hoja: str,
) -> LiquidacionDimanno:
    ruta = Path(ruta_archivo)

    if not ruta.is_file():
        raise FileNotFoundError(
            f"No existe el archivo: {ruta}"
        )

    coincidencia = PATRON_HOJA.match(nombre_hoja.strip())

    if not coincidencia:
        raise FormatoLiquidacionError(
            "El nombre de la hoja debe seguir el formato "
            "'FT 5292 W15'."
        )

    factura_corta = coincidencia.group("factura")
    semana = int(coincidencia.group("semana"))

    libro = load_workbook(
        filename=ruta,
        read_only=True,
        data_only=True,
        keep_links=True,
    )

    try:
        if nombre_hoja not in libro.sheetnames:
            raise FormatoLiquidacionError(
                f"No existe la hoja '{nombre_hoja}'."
            )

        hoja = libro[nombre_hoja]

        factura_celda = str(
            convertir_entero(
                hoja.cell(row=6, column=4).value,
                "número de factura",
            )
        ).zfill(4)

        if factura_celda != factura_corta:
            raise FormatoLiquidacionError(
                "La factura indicada en el nombre de la hoja "
                "no coincide con la factura dentro de la liquidación."
            )

        contenedores = separar_contenedores(
            hoja.cell(row=7, column=4).value
        )

        naviera = str(
            hoja.cell(row=8, column=1).value or ""
        ).strip().upper()

        destino = str(
            hoja.cell(row=9, column=1).value or ""
        ).strip().upper()

        productos, total_cajas, total_venta = extraer_productos(
            hoja
        )

        gastos, rubros_no_mapeados = extraer_gastos(hoja)

        return LiquidacionDimanno(
            archivo=ruta.name,
            hoja=nombre_hoja,
            factura_corta=factura_corta,
            semana=semana,
            contenedores=contenedores,
            naviera=naviera,
            destino=destino,
            total_cajas=total_cajas,
            total_venta_eur=total_venta,
            productos=productos,
            gastos=gastos,
            rubros_no_mapeados=rubros_no_mapeados,
        )

    finally:
        libro.close()


def convertir_a_json(valor: Any) -> Any:
    if isinstance(valor, Decimal):
        return format(valor, "f")

    if is_dataclass(valor):
        return convertir_a_json(asdict(valor))

    if isinstance(valor, dict):
        return {
            clave: convertir_a_json(contenido)
            for clave, contenido in valor.items()
        }

    if isinstance(valor, (list, tuple)):
        return [
            convertir_a_json(elemento)
            for elemento in valor
        ]

    return valor


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Prueba del extractor de liquidaciones Di Manno."
    )

    parser.add_argument(
        "--archivo",
        required=True,
        help="Ruta del archivo de liquidación.",
    )

    parser.add_argument(
        "--hoja",
        required=True,
        help="Hoja que se desea procesar. Ejemplo: FT 5292 W15.",
    )

    argumentos = parser.parse_args()

    liquidacion = extraer_liquidacion(
        ruta_archivo=argumentos.archivo,
        nombre_hoja=argumentos.hoja,
    )

    print(
        json.dumps(
            convertir_a_json(liquidacion),
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()