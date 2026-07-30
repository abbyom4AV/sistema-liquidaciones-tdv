from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass, is_dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from services.dimanno.matcher import (
    FormatoDespachosError,
    SinCoincidenciasError,
    convertir_entero,
    detectar_encabezados,
    interpretar_semana,
    normalizar_texto,
    obtener_factura_corta,
    texto_limpio,
)


CLIENTE_MASTER = "MASTERFRUITS"


class ErrorMatcherMaster(Exception):
    """Error general al cruzar Despachos para Master."""


class FormatoDespachosMasterError(ErrorMatcherMaster):
    """El archivo de Despachos no tiene el formato esperado."""


class SinCoincidenciasMasterError(ErrorMatcherMaster):
    """No hay líneas Master Fruits para la factura."""


@dataclass(frozen=True)
class LineaDespachoMaster:
    fila_excel: int
    semana: int
    anio: int
    semana_texto: str
    contenedor: str
    cliente: str
    barco: str
    puerto_destino: str
    tipo_empaque: str
    carton: str
    calibre: int
    total_cajas: int
    factura: str
    factura_corta: str

    @property
    def variante(self) -> str:
        carton_norm = normalizar_texto(self.carton)
        tipo_norm = normalizar_texto(self.tipo_empaque)

        if "VERTICAL" in carton_norm:
            return "VERTICAL"

        if tipo_norm == "ESPECIAL":
            return "ESPECIAL"

        return "VERDE"

    @property
    def tipo_fruta(self) -> str:
        if self.variante in {"VERTICAL", "ESPECIAL"}:
            return "ESPECIAL"
        return "VERDE"


@dataclass(frozen=True)
class ResultadoMatcherMaster:
    archivo: str
    hoja: str
    cliente_buscado: str
    factura_corta_buscada: str
    semana: int
    anio: int
    semana_texto: str
    lineas: tuple[LineaDespachoMaster, ...]
    total_cajas: int
    contenedores: tuple[str, ...]
    destinos: tuple[str, ...]


def _cliente_coincide(cliente_fila: str, buscado: str) -> bool:
    return normalizar_texto(cliente_fila).replace(
        " ",
        "",
    ) == normalizar_texto(buscado).replace(" ", "")


def buscar_lineas_despachos_master(
    ruta_archivo: str | Path,
    factura_corta: str,
    cliente: str = CLIENTE_MASTER,
) -> ResultadoMatcherMaster:
    ruta = Path(ruta_archivo)
    if not ruta.is_file():
        raise ErrorMatcherMaster(
            f"No existe el archivo de Despachos: {ruta}"
        )

    factura_objetivo = str(factura_corta).strip()
    if not re_es_factura_corta(factura_objetivo):
        raise FormatoDespachosMasterError(
            "La factura corta debe tener exactamente "
            "4 dígitos."
        )

    libro = load_workbook(
        ruta,
        read_only=True,
        data_only=True,
    )

    try:
        if "Base Datos" not in libro.sheetnames:
            raise FormatoDespachosMasterError(
                "No existe la hoja 'Base Datos' en Despachos."
            )

        hoja = libro["Base Datos"]

        try:
            fila_encabezado, posiciones = detectar_encabezados(
                hoja
            )
        except FormatoDespachosError as error:
            raise FormatoDespachosMasterError(str(error)) from error

        lineas: list[LineaDespachoMaster] = []

        for numero_fila, fila in enumerate(
            hoja.iter_rows(
                min_row=fila_encabezado + 1,
                values_only=True,
            ),
            start=fila_encabezado + 1,
        ):
            if fila is None or all(
                valor is None or str(valor).strip() == ""
                for valor in fila
            ):
                continue

            try:
                cliente_fila = texto_limpio(
                    _valor(fila, posiciones, "cliente")
                )
                if not _cliente_coincide(
                    cliente_fila,
                    cliente,
                ):
                    continue

                factura_valor = _valor(
                    fila,
                    posiciones,
                    "factura",
                )
                if (
                    factura_valor is None
                    or str(factura_valor).strip() == ""
                ):
                    continue

                factura_corta_fila = obtener_factura_corta(
                    factura_valor
                )
                if factura_corta_fila != factura_objetivo:
                    continue

                semana_valor = _valor(
                    fila,
                    posiciones,
                    "semana",
                )
                semana, anio_semana = interpretar_semana(
                    semana_valor
                )
                anio = convertir_entero(
                    _valor(fila, posiciones, "anio"),
                    "año",
                )
                if anio_semana is not None and anio_semana != anio:
                    # Preferir el año de la columna Año.
                    pass

                linea = LineaDespachoMaster(
                    fila_excel=numero_fila,
                    semana=semana,
                    anio=anio,
                    semana_texto=texto_limpio(semana_valor),
                    contenedor=texto_limpio(
                        _valor(
                            fila,
                            posiciones,
                            "contenedor",
                        )
                    ),
                    cliente=cliente_fila,
                    barco=texto_limpio(
                        _valor(fila, posiciones, "barco")
                    ),
                    puerto_destino=texto_limpio(
                        _valor(
                            fila,
                            posiciones,
                            "puerto_destino",
                        )
                    ),
                    tipo_empaque=texto_limpio(
                        _valor(
                            fila,
                            posiciones,
                            "tipo_empaque",
                        )
                    ),
                    carton=texto_limpio(
                        _valor(fila, posiciones, "carton")
                    ),
                    calibre=convertir_entero(
                        _valor(fila, posiciones, "calibre"),
                        "calibre",
                    ),
                    total_cajas=convertir_entero(
                        _valor(
                            fila,
                            posiciones,
                            "total_cajas",
                        ),
                        "total_cajas",
                    ),
                    factura=texto_limpio(factura_valor),
                    factura_corta=factura_corta_fila,
                )
            except (
                FormatoDespachosError,
                ValueError,
                TypeError,
            ) as error:
                raise FormatoDespachosMasterError(
                    f"Error en fila {numero_fila} de "
                    f"Despachos: {error}"
                ) from error

            lineas.append(linea)

        if not lineas:
            raise SinCoincidenciasMasterError(
                "No se encontraron líneas en Despachos para "
                f"cliente '{cliente}' y factura "
                f"'{factura_objetivo}'."
            )

        semanas = {(linea.semana, linea.anio) for linea in lineas}
        if len(semanas) != 1:
            detalle = ", ".join(
                f"W{semana}-{anio}"
                for semana, anio in sorted(semanas)
            )
            raise FormatoDespachosMasterError(
                "Las líneas coincidentes pertenecen a "
                f"varias semanas: {detalle}."
            )

        semana, anio = next(iter(semanas))
        semana_texto = lineas[0].semana_texto or (
            f"{semana:02d}-{anio}"
        )
        total_cajas = sum(linea.total_cajas for linea in lineas)
        contenedores = tuple(
            dict.fromkeys(
                linea.contenedor
                for linea in lineas
                if linea.contenedor
            )
        )
        destinos = tuple(
            dict.fromkeys(
                linea.puerto_destino.strip().upper()
                for linea in lineas
                if linea.puerto_destino.strip()
            )
        )

        return ResultadoMatcherMaster(
            archivo=ruta.name,
            hoja="Base Datos",
            cliente_buscado=cliente,
            factura_corta_buscada=factura_objetivo,
            semana=semana,
            anio=anio,
            semana_texto=semana_texto,
            lineas=tuple(lineas),
            total_cajas=total_cajas,
            contenedores=contenedores,
            destinos=destinos,
        )
    except SinCoincidenciasError as error:
        raise SinCoincidenciasMasterError(str(error)) from error
    finally:
        libro.close()


def re_es_factura_corta(valor: str) -> bool:
    return bool(valor) and valor.isdigit() and len(valor) == 4


def _valor(
    fila: tuple[Any, ...],
    posiciones: dict[str, int],
    campo: str,
) -> Any:
    indice = posiciones[campo] - 1
    if indice >= len(fila):
        return None
    return fila[indice]


def convertir_a_json(valor: Any) -> Any:
    if isinstance(valor, Decimal):
        return format(valor, "f")

    if is_dataclass(valor) and not isinstance(valor, type):
        return convertir_a_json(asdict(valor))

    if isinstance(valor, dict):
        return {
            clave: convertir_a_json(contenido)
            for clave, contenido in valor.items()
        }

    if isinstance(valor, (list, tuple)):
        return [
            convertir_a_json(elemento)
            for elemento in valor
        ]

    return valor


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Busca líneas de Despachos para Master Fruits."
        ),
    )
    parser.add_argument("--despachos", required=True)
    parser.add_argument("--factura", required=True)
    parser.add_argument(
        "--cliente",
        default=CLIENTE_MASTER,
    )
    argumentos = parser.parse_args()
    resultado = buscar_lineas_despachos_master(
        ruta_archivo=argumentos.despachos,
        factura_corta=argumentos.factura,
        cliente=argumentos.cliente,
    )
    print(json.dumps(convertir_a_json(resultado), indent=2))


if __name__ == "__main__":
    main()
