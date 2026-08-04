from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import asdict, dataclass, is_dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


COLUMNAS_GASTO = (
    "Deduction Eur",
    "Fresco Eur",
    "Transportation Eur",
    "Logistica Eur",
    "Handling Eur",
    "Transportation Cus. Eur",
)

# Orden importa: patrones más específicos primero.
MAPEO_GASTOS: tuple[tuple[tuple[str, ...], str], ...] = (
    (("COSIARMA", "THC/BL", "THC/BL FEE"), "Deduction Eur"),
    (("FRESCO-DELIVERY", "FRESCO DELIVERY"), "Fresco Eur"),
    (
        (
            "FRESCO-TRANSPORT",
            "FRESCO TRANSPORT",
            "TRANSPORT TO WEREHOUSE",
            "TRANSPORT TO WAREHOUSE",
            "TERMINAL COSTS",
        ),
        "Transportation Eur",
    ),
    (
        (
            "LOGISTICA-WEREHOUSE",
            "LOGISTICA-WAREHOUSE",
            "WAREHOUSE DESCHARGED",
            "WEREHOUSE DESCHARGED",
        ),
        "Logistica Eur",
    ),
    (
        ("LOGISTICA-HANDLING", "HANDLING AND PICKING"),
        "Handling Eur",
    ),
    (
        (
            "CAA-TRANSPORT",
            "TRANSPORT TO CUSTOMER",
            "TRANSPORTATION CUSTOMER",
        ),
        "Transportation Cus. Eur",
    ),
)

_CONTENEDOR_RE = re.compile(r"\b([A-Z]{4}\d{7})\b", re.IGNORECASE)
_VERTICAL_RE = re.compile(
    r"^\s*(\d+)\s+VERTICAL\s*$",
    re.IGNORECASE,
)
_FACTURA_RE = re.compile(r"(\d{3,})")
_COMMISSION_RE = re.compile(
    r"commission\s+"
    r"(NO|(?:\d+(?:[.,]\d+)?\s*%))"
    r"\s*(?:n[°ºo.]?\s*)?"
    r"([A-Z]{4}\d{7})?",
    re.IGNORECASE,
)


class ErrorExtraccionSifa(Exception):
    """Error al leer la liquidación Excel de SIFA."""


@dataclass(frozen=True)
class LineaProductoSifa:
    contenedor: str
    marca: str
    carton: str
    calibre: int
    total_cajas: int
    notes: str
    tipo_fruta: str
    precio_eur: Decimal
    amount_eur: Decimal
    sin_comision: bool
    fila_excel: int


@dataclass(frozen=True)
class ComisionContenedorSifa:
    contenedor: str
    monto_eur: Decimal
    sin_comision: bool
    texto_origen: str


@dataclass(frozen=True)
class TotalContenedorSifa:
    contenedor: str
    total_cajas: int
    total_venta_eur: Decimal


@dataclass(frozen=True)
class LiquidacionSifa:
    archivo: str
    hoja: str
    vessel: str
    destino_header: str
    factura: str
    factura_corta: str
    orden: str
    contenedores_header: tuple[str, ...]
    lineas: tuple[LineaProductoSifa, ...]
    gastos: dict[str, Decimal]
    rubros_mapeados: tuple[str, ...]
    rubros_no_mapeados: tuple[str, ...]
    comisiones_contenedor: tuple[ComisionContenedorSifa, ...]
    totales_contenedor: tuple[TotalContenedorSifa, ...]
    total_costos_excel: Decimal | None
    total_cajas: int
    total_venta_eur: Decimal


def _a_decimal(valor: Any) -> Decimal | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, Decimal):
        return valor
    if isinstance(valor, (int, float)):
        return Decimal(str(valor))
    texto = str(valor).strip()
    if not texto:
        return None
    texto = (
        texto.replace("€", "")
        .replace("EUR", "")
        .replace(" ", "")
        .strip()
    )
    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        partes = texto.split(",")
        if len(partes) == 2 and len(partes[1]) <= 2:
            texto = texto.replace(",", ".")
        else:
            texto = texto.replace(",", "")
    try:
        return Decimal(texto)
    except (InvalidOperation, ValueError):
        return None


def _texto(valor: Any) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def _normalizar_gasto(texto: str) -> str:
    return re.sub(r"\s+", " ", texto.upper()).strip()


def _mapear_columna_gasto(etiqueta: str) -> str | None:
    n = _normalizar_gasto(etiqueta)
    if not n or n.startswith("DETALLE") or n == "DEDUCTIONS:":
        return None
    if "GASTOS TOTAL" in n or n.startswith("TOTAL OF COST"):
        return None
    if n.startswith("COMMISSION"):
        return None
    for patrones, columna in MAPEO_GASTOS:
        if any(p in n for p in patrones):
            return columna
    return None


def _es_encabezado_bloque(valores: list[Any]) -> bool:
    textos = {_texto(v).upper() for v in valores if _texto(v)}
    return "MARCA" in textos and "COUNT" in textos and (
        "CARTONS" in textos or "CARTON" in textos
    )


def _contenedor_de_celda(valor: Any) -> str | None:
    texto = _texto(valor).upper().replace(" ", "")
    if not texto:
        return None
    match = _CONTENEDOR_RE.fullmatch(texto)
    if match:
        return match.group(1).upper()
    # A veces el encabezado es solo el contenedor.
    match = _CONTENEDOR_RE.search(_texto(valor).upper())
    if match and len(texto) <= 20:
        return match.group(1).upper()
    return None


def _parsear_calibre_y_carton(
    count_valor: Any,
    marca: str,
) -> tuple[int, str]:
    marca_limpia = re.sub(r"\s+", " ", marca).strip()
    if isinstance(count_valor, (int, float, Decimal)):
        return int(count_valor), marca_limpia

    texto = _texto(count_valor)
    match = _VERTICAL_RE.match(texto)
    if match:
        calibre = int(match.group(1))
        carton = f"VERTICAL {marca_limpia}".strip()
        return calibre, carton

    solo = re.match(r"^\s*(\d+)\s*$", texto)
    if solo:
        return int(solo.group(1)), marca_limpia

    raise ErrorExtraccionSifa(
        f"No se pudo interpretar COUNT={count_valor!r} "
        f"para marca {marca!r}."
    )


def _tipo_fruta_desde_notes(notes: str) -> str:
    n = notes.strip().upper()
    if not n:
        return "ESPECIAL"
    if "VERDE" in n:
        return "VERDE"
    return "ESPECIAL"


def _extraer_factura_corta(valor: Any) -> str:
    texto = _texto(valor)
    if not texto:
        return ""
    # "5513 DEL 25/05/2026" o solo "5513"
    match = re.search(
        r"(?:INV\.?\s*)?(\d{3,})\b",
        texto,
        re.IGNORECASE,
    )
    if match:
        return match.group(1)[-4:] if len(match.group(1)) > 4 else match.group(1)
    digitos = "".join(ch for ch in texto if ch.isdigit())
    return digitos[-4:] if digitos else ""


def _monto_en_fila(valores: list[Any], desde: int = 0) -> Decimal | None:
    for valor in valores[desde:]:
        monto = _a_decimal(valor)
        if monto is not None:
            return monto
    return None


def extraer_liquidacion_sifa(
    ruta_archivo: str | Path,
) -> LiquidacionSifa:
    ruta = Path(ruta_archivo)
    if not ruta.is_file():
        raise ErrorExtraccionSifa(
            f"No existe el archivo de liquidación: {ruta}"
        )

    libro = load_workbook(ruta, data_only=True, read_only=True)
    try:
        nombre_hoja = libro.sheetnames[0]
        hoja = libro[nombre_hoja]
        filas = list(
            hoja.iter_rows(min_row=1, max_col=15, values_only=True)
        )
    finally:
        libro.close()

    vessel = ""
    destino_header = ""
    factura = ""
    factura_corta = ""
    orden = ""
    contenedores_header: list[str] = []

    lineas: list[LineaProductoSifa] = []
    gastos: dict[str, Decimal] = {
        col: Decimal("0") for col in COLUMNAS_GASTO
    }
    rubros_mapeados: list[str] = []
    rubros_no_mapeados: list[str] = []
    comisiones: list[ComisionContenedorSifa] = []
    totales_contenedor: list[TotalContenedorSifa] = []
    total_costos: Decimal | None = None

    contenedor_actual = ""
    en_bloque_productos = False
    cajas_bloque = 0
    venta_bloque = Decimal("0")

    for idx, fila in enumerate(filas, start=1):
        valores = list(fila or ())
        while len(valores) < 12:
            valores.append(None)

        c3 = _texto(valores[2])
        c4 = _texto(valores[3])
        c7 = _texto(valores[6])
        c8 = _texto(valores[7])
        unidos = " ".join(
            _texto(v) for v in valores if _texto(v)
        )

        if not vessel and "VESSEL" in unidos.upper():
            for valor in valores:
                t = _texto(valor)
                if "VESSEL" in t.upper():
                    vessel = re.sub(
                        r"(?i)^VESSEL:\s*",
                        "",
                        t,
                    ).strip()
                    break
            # Destino a la derecha de "at"
            for i, valor in enumerate(valores):
                if _texto(valor).upper() == "AT" and i + 1 < len(
                    valores
                ):
                    destino_header = _texto(valores[i + 1])
                    break
            if not destino_header:
                for valor in valores:
                    t = _texto(valor)
                    if "ITALY" in t.upper() or "LIGURE" in t.upper():
                        destino_header = t
                        break

        if c7.upper().startswith("INV") or (
            c7.upper() == "INV." or c7.upper() == "INV"
        ):
            factura = c8 or factura
            factura_corta = _extraer_factura_corta(factura)

        if "ORDEN" in c7.upper() and c8:
            orden = c8

        if c3.upper().startswith("CONTAINER"):
            for valor in valores[3:]:
                for match in _CONTENEDOR_RE.finditer(
                    _texto(valor).upper()
                ):
                    cont = match.group(1).upper()
                    if cont not in contenedores_header:
                        contenedores_header.append(cont)

        # Continuación de lista de contenedores en filas siguientes.
        if not en_bloque_productos and "/" in c3 and _CONTENEDOR_RE.search(
            c3.upper()
        ):
            for match in _CONTENEDOR_RE.finditer(c3.upper()):
                cont = match.group(1).upper()
                if cont not in contenedores_header:
                    contenedores_header.append(cont)

        if _es_encabezado_bloque(valores):
            cont = _contenedor_de_celda(valores[2])
            if cont:
                contenedor_actual = cont
                if cont not in contenedores_header:
                    contenedores_header.append(cont)
            en_bloque_productos = True
            cajas_bloque = 0
            venta_bloque = Decimal("0")
            continue

        if en_bloque_productos:
            etiqueta = c3.upper()
            if etiqueta == "TOTAL":
                cajas_excel = _a_decimal(valores[5])
                venta_excel = _a_decimal(valores[7])
                cajas_final = (
                    int(cajas_excel)
                    if cajas_excel is not None
                    else cajas_bloque
                )
                venta_final = (
                    venta_excel
                    if venta_excel is not None
                    else venta_bloque
                )
                if contenedor_actual:
                    totales_contenedor.append(
                        TotalContenedorSifa(
                            contenedor=contenedor_actual,
                            total_cajas=cajas_final,
                            total_venta_eur=venta_final,
                        )
                    )
                en_bloque_productos = False
                contenedor_actual = ""
                cajas_bloque = 0
                venta_bloque = Decimal("0")
                continue
            if etiqueta == "PINES":
                if not contenedor_actual:
                    raise ErrorExtraccionSifa(
                        f"Línea Pines sin contenedor (fila {idx})."
                    )
                marca = c4
                try:
                    calibre, carton = _parsear_calibre_y_carton(
                        valores[4],
                        marca,
                    )
                except ErrorExtraccionSifa as error:
                    raise ErrorExtraccionSifa(
                        f"Fila {idx}: {error}"
                    ) from error
                cajas = _a_decimal(valores[5])
                if cajas is None:
                    raise ErrorExtraccionSifa(
                        f"Fila {idx}: CARTONS inválido."
                    )
                notes = _texto(valores[6])
                amount = _a_decimal(valores[7]) or Decimal("0")
                precio = _a_decimal(valores[8])
                if precio is None:
                    raise ErrorExtraccionSifa(
                        f"Fila {idx}: falta €/CRT/NET."
                    )
                flag = _texto(valores[9]).upper()
                notes_n = notes.upper()
                # CODIGO en NOTES ⇒ sin comisión (a veces no ponen
                # NO COMMISSION a la par). También sin comisión si
                # la columna lateral lo indica explícitamente.
                sin_comision = (
                    "NO COMMISSION" in flag
                    or "CODIGO" in notes_n
                )
                cajas_bloque += int(cajas)
                venta_bloque += amount
                lineas.append(
                    LineaProductoSifa(
                        contenedor=contenedor_actual,
                        marca=marca,
                        carton=carton,
                        calibre=calibre,
                        total_cajas=int(cajas),
                        notes=notes,
                        tipo_fruta=_tipo_fruta_desde_notes(notes),
                        precio_eur=precio,
                        amount_eur=amount,
                        sin_comision=sin_comision,
                        fila_excel=idx,
                    )
                )
                continue
            # Otro encabezado de contenedor sin pasar por TOTAL.
            cont = _contenedor_de_celda(valores[2])
            if cont and _es_encabezado_bloque(valores):
                contenedor_actual = cont
                cajas_bloque = 0
                venta_bloque = Decimal("0")
                continue

        # Gastos / comisiones (fuera de bloques de producto).
        etiqueta = c3
        etiqueta_n = _normalizar_gasto(etiqueta)
        if not etiqueta_n:
            continue

        if "TOTAL OF COST" in etiqueta_n:
            total_costos = _monto_en_fila(valores, desde=3)
            continue

        if etiqueta_n.startswith("COMMISSION"):
            match = _COMMISSION_RE.search(unidos)
            cont = ""
            if match and match.group(2):
                cont = match.group(2).upper()
            else:
                found = _CONTENEDOR_RE.search(unidos.upper())
                if found:
                    cont = found.group(1).upper()
            if not cont:
                continue
            token = (match.group(1) if match else "").strip().upper()
            es_no = token == "NO" or re.search(
                r"\bNO\b",
                etiqueta_n,
            )
            if es_no:
                comisiones.append(
                    ComisionContenedorSifa(
                        contenedor=cont,
                        monto_eur=Decimal("0"),
                        sin_comision=True,
                        texto_origen=unidos,
                    )
                )
            else:
                # El % no es el monto; el euro viene en la fila.
                monto = _a_decimal(valores[7])
                if monto is None:
                    monto = _a_decimal(valores[8])
                if monto is None:
                    monto = _monto_en_fila(valores, desde=3) or (
                        Decimal("0")
                    )
                comisiones.append(
                    ComisionContenedorSifa(
                        contenedor=cont,
                        monto_eur=monto,
                        sin_comision=False,
                        texto_origen=unidos,
                    )
                )
            continue

        columna = _mapear_columna_gasto(etiqueta)
        if columna is None:
            # Solo avisar por rubros principales (monto en col amount),
            # no por líneas de Detalle.
            monto_principal = _a_decimal(valores[7])
            if monto_principal is not None and monto_principal != 0:
                if any(
                    k in etiqueta_n
                    for k in (
                        "FEE",
                        "COST",
                        "TRANSPORT",
                        "LOGIST",
                        "FRESCO",
                        "HANDLING",
                        "CAA",
                        "COSIARMA",
                        "DEDUCTION",
                    )
                ) and not etiqueta_n.startswith("DETALLE"):
                    rubros_no_mapeados.append(
                        f"{etiqueta}={monto_principal}"
                    )
            continue

        # Totales de rubro van en la columna AMOUNT (índice 7).
        monto = _a_decimal(valores[7])
        if monto is None:
            monto = _monto_en_fila(valores, desde=3)
        if monto is None:
            continue
        gastos[columna] = gastos.get(columna, Decimal("0")) + monto
        rubros_mapeados.append(f"{etiqueta}→{columna}={monto}")

    if not lineas:
        raise ErrorExtraccionSifa(
            "No se encontraron líneas 'Pines' en la liquidación."
        )

    contenedores_lineas = tuple(
        dict.fromkeys(ln.contenedor for ln in lineas)
    )
    if not contenedores_header:
        contenedores_header = list(contenedores_lineas)

    # Si algún bloque no trajo fila TOTAL, armar desde líneas.
    vistos = {
        t.contenedor.replace(" ", "").upper()
        for t in totales_contenedor
    }
    cajas_por: dict[str, int] = defaultdict(int)
    venta_por: dict[str, Decimal] = defaultdict(
        lambda: Decimal("0")
    )
    for ln in lineas:
        cajas_por[ln.contenedor] += ln.total_cajas
        venta_por[ln.contenedor] += ln.amount_eur
    for cont, cajas in cajas_por.items():
        clave = cont.replace(" ", "").upper()
        if clave in vistos:
            continue
        totales_contenedor.append(
            TotalContenedorSifa(
                contenedor=cont,
                total_cajas=cajas,
                total_venta_eur=venta_por[cont],
            )
        )

    return LiquidacionSifa(
        archivo=ruta.name,
        hoja=nombre_hoja,
        vessel=vessel,
        destino_header=destino_header,
        factura=factura,
        factura_corta=factura_corta or _extraer_factura_corta(
            factura
        ),
        orden=orden,
        contenedores_header=tuple(contenedores_header),
        lineas=tuple(lineas),
        gastos=gastos,
        rubros_mapeados=tuple(rubros_mapeados),
        rubros_no_mapeados=tuple(rubros_no_mapeados),
        comisiones_contenedor=tuple(comisiones),
        totales_contenedor=tuple(totales_contenedor),
        total_costos_excel=total_costos,
        total_cajas=sum(ln.total_cajas for ln in lineas),
        total_venta_eur=sum(
            (ln.amount_eur for ln in lineas),
            Decimal("0"),
        ),
    )


def convertir_a_json(valor: Any) -> Any:
    if isinstance(valor, Decimal):
        return format(valor, "f")
    if is_dataclass(valor) and not isinstance(valor, type):
        return convertir_a_json(asdict(valor))
    if isinstance(valor, dict):
        return {k: convertir_a_json(v) for k, v in valor.items()}
    if isinstance(valor, (list, tuple)):
        return [convertir_a_json(v) for v in valor]
    return valor


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extrae liquidación Excel SIFA.",
    )
    parser.add_argument("excel")
    args = parser.parse_args()
    print(
        json.dumps(
            convertir_a_json(extraer_liquidacion_sifa(args.excel)),
            indent=2,
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
