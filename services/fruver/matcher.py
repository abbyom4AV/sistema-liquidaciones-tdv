from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass, is_dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from services.dimanno.matcher import (
    FormatoDespachosError,
    convertir_entero,
    detectar_encabezados,
    interpretar_semana,
    obtener_factura_corta,
    texto_limpio,
)
from services.fruver.extractor import normalizar_destino, normalizar_texto


CLIENTE_FRUVER = "FRU&VER"
CLIENTE_FRUVER_ALIASES = (
    "FRU&VER",
    "FRUVER",
    "FRU VER",
    "FRU Y VER",
    "CIA FRU&VER",
    "CIA. FRU&VER",
    "FRU&VER MADRID",
)


class ErrorMatcherFruver(Exception):
    """Error general al cruzar Despachos para FRU&VER."""


class FormatoDespachosFruverError(ErrorMatcherFruver):
    """El archivo de Despachos no tiene el formato esperado."""


class SinCoincidenciasFruverError(ErrorMatcherFruver):
    """No hay líneas FRU&VER para los criterios."""


@dataclass(frozen=True)
class LineaDespachoFruver:
    fila_excel: int
    semana: int
    anio: int
    semana_texto: str
    contenedor: str
    cliente: str
    barco: str
    puerto_destino: str
    tipo_empaque: str
    carton: str
    calibre: int
    total_cajas: int
    factura: str
    factura_corta: str


@dataclass(frozen=True)
class ResultadoMatcherFruver:
    archivo: str
    hoja: str
    cliente_buscado: str
    semana: int
    anio: int
    destino_buscado: str
    factura_corta_buscada: str
    semana_texto: str
    lineas: tuple[LineaDespachoFruver, ...]
    total_cajas: int
    contenedores: tuple[str, ...]
    destinos: tuple[str, ...]
    naves: tuple[str, ...]


def _cliente_coincide(cliente_fila: str, buscado: str) -> bool:
    a = normalizar_texto(cliente_fila).replace(" ", "").replace("&", "")
    b = normalizar_texto(buscado).replace(" ", "").replace("&", "")
    if not a:
        return False
    if a == b or a.startswith("FRUVER") or "FRUVER" in a:
        return True
    return a.replace("Y", "") == "FRUVER" or a.startswith("FRUYVER")


def _destino_coincide(destino_fila: str, buscado: str) -> bool:
    a = normalizar_destino(destino_fila)
    b = normalizar_destino(buscado)
    if not a or not b:
        return False
    return a == b or a in b or b in a


def _valor(
    fila: tuple[Any, ...],
    posiciones: dict[str, int],
    campo: str,
) -> Any:
    indice = posiciones[campo] - 1
    if indice >= len(fila):
        return None
    return fila[indice]


def buscar_lineas_despachos_fruver(
    ruta_archivo: str | Path,
    semana: int,
    anio: int,
    destino: str,
    factura_corta: str,
    cliente: str = CLIENTE_FRUVER,
) -> ResultadoMatcherFruver:
    ruta = Path(ruta_archivo)
    if not ruta.is_file():
        raise ErrorMatcherFruver(
            f"No existe el archivo de Despachos: {ruta}"
        )
    if not (1 <= int(semana) <= 53):
        raise FormatoDespachosFruverError(
            "La semana debe estar entre 1 y 53."
        )
    factura_objetivo = str(factura_corta or "").strip()
    if len(factura_objetivo) != 4 or not factura_objetivo.isdigit():
        raise FormatoDespachosFruverError(
            "La factura corta debe tener exactamente 4 dígitos."
        )

    libro = load_workbook(ruta, read_only=True, data_only=True)
    try:
        if "Base Datos" not in libro.sheetnames:
            raise FormatoDespachosFruverError(
                "No existe la hoja 'Base Datos' en Despachos."
            )
        hoja = libro["Base Datos"]
        try:
            fila_encabezado, posiciones = detectar_encabezados(hoja)
        except FormatoDespachosError as error:
            raise FormatoDespachosFruverError(str(error)) from error

        lineas: list[LineaDespachoFruver] = []
        for numero_fila, fila in enumerate(
            hoja.iter_rows(
                min_row=fila_encabezado + 1,
                values_only=True,
            ),
            start=fila_encabezado + 1,
        ):
            if fila is None or all(
                v is None or str(v).strip() == "" for v in fila
            ):
                continue
            try:
                cliente_fila = texto_limpio(
                    _valor(fila, posiciones, "cliente")
                )
                if not _cliente_coincide(cliente_fila, cliente):
                    continue
                semana_valor = _valor(fila, posiciones, "semana")
                semana_fila, _ = interpretar_semana(semana_valor)
                anio_fila = convertir_entero(
                    _valor(fila, posiciones, "anio"),
                    "año",
                )
                if semana_fila != int(semana) or anio_fila != int(anio):
                    continue
                factura_valor = _valor(fila, posiciones, "factura")
                factura_txt = texto_limpio(factura_valor)
                factura_fila = ""
                if factura_txt:
                    try:
                        factura_fila = obtener_factura_corta(
                            factura_valor
                        )
                    except Exception:
                        factura_fila = ""
                if factura_fila != factura_objetivo:
                    continue
                puerto = texto_limpio(
                    _valor(fila, posiciones, "puerto_destino")
                )
                if not _destino_coincide(puerto, destino):
                    continue
                lineas.append(
                    LineaDespachoFruver(
                        fila_excel=numero_fila,
                        semana=semana_fila,
                        anio=anio_fila,
                        semana_texto=texto_limpio(semana_valor),
                        contenedor=texto_limpio(
                            _valor(fila, posiciones, "contenedor")
                        ).upper(),
                        cliente=cliente_fila,
                        barco=texto_limpio(
                            _valor(fila, posiciones, "barco")
                        ),
                        puerto_destino=puerto,
                        tipo_empaque=texto_limpio(
                            _valor(fila, posiciones, "tipo_empaque")
                        ),
                        carton=texto_limpio(
                            _valor(fila, posiciones, "carton")
                        ),
                        calibre=convertir_entero(
                            _valor(fila, posiciones, "calibre"),
                            "calibre",
                        ),
                        total_cajas=convertir_entero(
                            _valor(fila, posiciones, "total_cajas"),
                            "total_cajas",
                        ),
                        factura=factura_txt,
                        factura_corta=factura_fila,
                    )
                )
            except (FormatoDespachosError, ValueError, TypeError) as error:
                raise FormatoDespachosFruverError(
                    f"Error en fila {numero_fila}: {error}"
                ) from error

        if not lineas:
            raise SinCoincidenciasFruverError(
                "No se encontraron líneas en Despachos para "
                f"FRU&VER, semana {semana}, año {anio}, "
                f"destino '{destino}', factura {factura_objetivo}."
            )

        destinos = tuple(
            dict.fromkeys(
                ln.puerto_destino.strip().upper()
                for ln in lineas
                if ln.puerto_destino.strip()
            )
        )
        contenedores = tuple(
            dict.fromkeys(ln.contenedor for ln in lineas if ln.contenedor)
        )
        naves = tuple(
            dict.fromkeys(ln.barco for ln in lineas if ln.barco)
        )
        semana_texto = lineas[0].semana_texto or f"{semana:02d}-{anio}"
        return ResultadoMatcherFruver(
            archivo=ruta.name,
            hoja="Base Datos",
            cliente_buscado=cliente,
            semana=int(semana),
            anio=int(anio),
            destino_buscado=destino.strip(),
            factura_corta_buscada=factura_objetivo,
            semana_texto=semana_texto,
            lineas=tuple(lineas),
            total_cajas=sum(ln.total_cajas for ln in lineas),
            contenedores=contenedores,
            destinos=destinos,
            naves=naves,
        )
    finally:
        libro.close()


def convertir_a_json(valor: Any) -> Any:
    if isinstance(valor, Decimal):
        return format(valor, "f")
    if is_dataclass(valor) and not isinstance(valor, type):
        return convertir_a_json(asdict(valor))
    if isinstance(valor, dict):
        return {k: convertir_a_json(v) for k, v in valor.items()}
    if isinstance(valor, (list, tuple)):
        return [convertir_a_json(v) for v in valor]
    return valor


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--despachos", required=True)
    parser.add_argument("--semana", type=int, required=True)
    parser.add_argument("--anio", type=int, required=True)
    parser.add_argument("--destino", required=True)
    parser.add_argument("--factura", required=True)
    args = parser.parse_args()
    print(
        json.dumps(
            convertir_a_json(
                buscar_lineas_despachos_fruver(
                    args.despachos,
                    args.semana,
                    args.anio,
                    args.destino,
                    args.factura,
                )
            ),
            indent=2,
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
