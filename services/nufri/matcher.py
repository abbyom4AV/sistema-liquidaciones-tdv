from __future__ import annotations

import argparse
import json
import re
from dataclasses import asdict, dataclass, is_dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from services.dimanno.matcher import (
    FormatoDespachosError,
    SinCoincidenciasError,
    convertir_entero,
    detectar_encabezados,
    interpretar_semana,
    normalizar_texto,
    obtener_factura_corta,
    texto_limpio,
)


CLIENTE_NUFRI = "NUFRI"


class ErrorMatcherNufri(Exception):
    """Error general al cruzar Despachos para NUFRI."""


class FormatoDespachosNufriError(ErrorMatcherNufri):
    """El archivo de Despachos no tiene el formato esperado."""


class SinCoincidenciasNufriError(ErrorMatcherNufri):
    """No hay líneas NUFRI para los criterios."""


@dataclass(frozen=True)
class LineaDespachoNufri:
    fila_excel: int
    semana: int
    anio: int
    semana_texto: str
    contenedor: str
    cliente: str
    barco: str
    puerto_destino: str
    tipo_empaque: str
    carton: str
    calibre: int
    total_cajas: int
    factura: str
    factura_corta: str


@dataclass(frozen=True)
class ResultadoMatcherNufri:
    archivo: str
    hoja: str
    cliente_buscado: str
    factura_corta_buscada: str
    semana: int
    anio: int
    destino_buscado: str
    semana_texto: str
    lineas: tuple[LineaDespachoNufri, ...]
    total_cajas: int
    contenedores: tuple[str, ...]
    destinos: tuple[str, ...]


def _cliente_coincide(cliente_fila: str, buscado: str) -> bool:
    return normalizar_texto(cliente_fila).replace(
        " ",
        "",
    ) == normalizar_texto(buscado).replace(" ", "")


def _destino_coincide(destino_fila: str, buscado: str) -> bool:
    a = normalizar_texto(destino_fila)
    b = normalizar_texto(buscado)
    if not a or not b:
        return False
    return a == b or a in b or b in a


def re_es_factura_corta(valor: str) -> bool:
    return bool(valor) and valor.isdigit() and len(valor) == 4


def _valor(
    fila: tuple[Any, ...],
    posiciones: dict[str, int],
    campo: str,
) -> Any:
    indice = posiciones[campo] - 1
    if indice >= len(fila):
        return None
    return fila[indice]


def buscar_lineas_despachos_nufri(
    ruta_archivo: str | Path,
    *,
    semana: int,
    anio: int,
    destino: str,
    factura_corta: str,
    cliente: str = CLIENTE_NUFRI,
) -> ResultadoMatcherNufri:
    ruta = Path(ruta_archivo)
    if not ruta.is_file():
        raise ErrorMatcherNufri(
            f"No existe el archivo de Despachos: {ruta}"
        )
    if not (1 <= int(semana) <= 53):
        raise FormatoDespachosNufriError(
            "La semana debe estar entre 1 y 53."
        )

    factura_objetivo = str(factura_corta).strip()
    if not re_es_factura_corta(factura_objetivo):
        raise FormatoDespachosNufriError(
            "La factura corta debe tener exactamente 4 dígitos."
        )

    destino_obj = str(destino).strip()
    if not destino_obj:
        raise FormatoDespachosNufriError(
            "Indique el destino."
        )

    libro = load_workbook(
        ruta,
        read_only=True,
        data_only=True,
    )

    try:
        if "Base Datos" not in libro.sheetnames:
            raise FormatoDespachosNufriError(
                "No existe la hoja 'Base Datos' en Despachos."
            )

        hoja = libro["Base Datos"]

        try:
            fila_encabezado, posiciones = detectar_encabezados(
                hoja
            )
        except FormatoDespachosError as error:
            raise FormatoDespachosNufriError(str(error)) from error

        lineas: list[LineaDespachoNufri] = []

        for numero_fila, fila in enumerate(
            hoja.iter_rows(
                min_row=fila_encabezado + 1,
                values_only=True,
            ),
            start=fila_encabezado + 1,
        ):
            if fila is None or all(
                valor is None or str(valor).strip() == ""
                for valor in fila
            ):
                continue

            try:
                cliente_fila = texto_limpio(
                    _valor(fila, posiciones, "cliente")
                )
                if not _cliente_coincide(cliente_fila, cliente):
                    continue

                semana_valor = _valor(fila, posiciones, "semana")
                semana_fila, _ = interpretar_semana(semana_valor)
                anio_fila = convertir_entero(
                    _valor(fila, posiciones, "anio"),
                    "año",
                )
                if (
                    semana_fila != int(semana)
                    or anio_fila != int(anio)
                ):
                    continue

                puerto = texto_limpio(
                    _valor(fila, posiciones, "puerto_destino")
                )
                if not _destino_coincide(puerto, destino_obj):
                    continue

                factura_valor = _valor(
                    fila,
                    posiciones,
                    "factura",
                )
                if (
                    factura_valor is None
                    or str(factura_valor).strip() == ""
                ):
                    continue

                factura_corta_fila = obtener_factura_corta(
                    factura_valor
                )
                if factura_corta_fila != factura_objetivo:
                    continue

                linea = LineaDespachoNufri(
                    fila_excel=numero_fila,
                    semana=semana_fila,
                    anio=anio_fila,
                    semana_texto=texto_limpio(semana_valor),
                    contenedor=texto_limpio(
                        _valor(fila, posiciones, "contenedor")
                    ).upper(),
                    cliente=cliente_fila,
                    barco=texto_limpio(
                        _valor(fila, posiciones, "barco")
                    ),
                    puerto_destino=puerto,
                    tipo_empaque=texto_limpio(
                        _valor(
                            fila,
                            posiciones,
                            "tipo_empaque",
                        )
                    ),
                    carton=texto_limpio(
                        _valor(fila, posiciones, "carton")
                    ),
                    calibre=convertir_entero(
                        _valor(fila, posiciones, "calibre"),
                        "calibre",
                    ),
                    total_cajas=convertir_entero(
                        _valor(
                            fila,
                            posiciones,
                            "total_cajas",
                        ),
                        "total_cajas",
                    ),
                    factura=texto_limpio(factura_valor),
                    factura_corta=factura_corta_fila,
                )
            except (
                FormatoDespachosError,
                ValueError,
                TypeError,
            ) as error:
                raise FormatoDespachosNufriError(
                    f"Error en fila {numero_fila} de "
                    f"Despachos: {error}"
                ) from error

            lineas.append(linea)

        if not lineas:
            raise SinCoincidenciasNufriError(
                "No se encontraron líneas en Despachos para "
                f"{cliente}, semana {semana}, año {anio}, "
                f"destino '{destino_obj}', factura "
                f"'{factura_objetivo}'."
            )

        semana_texto = lineas[0].semana_texto or (
            f"{int(semana):02d}-{int(anio)}"
        )
        total_cajas = sum(linea.total_cajas for linea in lineas)
        contenedores = tuple(
            dict.fromkeys(
                linea.contenedor
                for linea in lineas
                if linea.contenedor
            )
        )
        destinos = tuple(
            dict.fromkeys(
                linea.puerto_destino.strip().upper()
                for linea in lineas
                if linea.puerto_destino.strip()
            )
        )

        return ResultadoMatcherNufri(
            archivo=ruta.name,
            hoja="Base Datos",
            cliente_buscado=cliente,
            factura_corta_buscada=factura_objetivo,
            semana=int(semana),
            anio=int(anio),
            destino_buscado=destino_obj,
            semana_texto=semana_texto,
            lineas=tuple(lineas),
            total_cajas=total_cajas,
            contenedores=contenedores,
            destinos=destinos,
        )
    except SinCoincidenciasError as error:
        raise SinCoincidenciasNufriError(str(error)) from error
    finally:
        libro.close()


def convertir_a_json(valor: Any) -> Any:
    if isinstance(valor, Decimal):
        return format(valor, "f")
    if is_dataclass(valor) and not isinstance(valor, type):
        return convertir_a_json(asdict(valor))
    if isinstance(valor, dict):
        return {
            clave: convertir_a_json(contenido)
            for clave, contenido in valor.items()
        }
    if isinstance(valor, (list, tuple)):
        return [convertir_a_json(elemento) for elemento in valor]
    return valor


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Busca líneas Despachos NUFRI.",
    )
    parser.add_argument("--despachos", required=True)
    parser.add_argument("--semana", type=int, required=True)
    parser.add_argument("--anio", type=int, required=True)
    parser.add_argument("--destino", required=True)
    parser.add_argument("--factura", required=True)
    parser.add_argument("--cliente", default=CLIENTE_NUFRI)
    args = parser.parse_args()
    print(
        json.dumps(
            convertir_a_json(
                buscar_lineas_despachos_nufri(
                    ruta_archivo=args.despachos,
                    semana=args.semana,
                    anio=args.anio,
                    destino=args.destino,
                    factura_corta=args.factura,
                    cliente=args.cliente,
                )
            ),
            indent=2,
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
